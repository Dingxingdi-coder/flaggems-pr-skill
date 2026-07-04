#!/usr/bin/env python3
import argparse
import os
import sys

import openpyxl

DEFAULT_NORM_XLSX = os.environ.get("FLAGGEMS_NORM_XLSX", "/data/dxd/规范名.xlsx")


def clean(value):
    if value is None:
        return ""
    return str(value).replace("\r", "").replace("_x000d_", "").strip()


def clean_op(name):
    return clean(name).replace("aten::", "", 1)


def find_columns(ws, need_path=False):
    headers = [clean(cell.value) for cell in ws[1]]

    try:
        name_col = headers.index("算子名") + 1
    except ValueError:
        raise SystemExit("missing column: 算子名")

    norm_col = None
    path_col = None
    for i, header in enumerate(headers, start=1):
        if header == "算子名（规范命名）" or "规范命名" in header:
            norm_col = i
        if header == "代码路径" or "代码路径" in header:
            path_col = i

    if norm_col is None:
        raise SystemExit("missing column: 算子名（规范命名）")
    if need_path and path_col is None:
        raise SystemExit("missing column: 代码路径")

    return name_col, norm_col, path_col


def lookup_norm_name(op_name, xlsx_path):
    op_name = clean_op(op_name)
    if not op_name:
        raise SystemExit("operator name is empty")
    if not os.path.isfile(xlsx_path):
        raise SystemExit(f"file not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    name_col, norm_col, _ = find_columns(ws)

    for row in ws.iter_rows(min_row=2):
        name = clean_op(row[name_col - 1].value)
        if name == op_name:
            norm_name = clean(row[norm_col - 1].value)
            if not norm_name:
                raise SystemExit(f"empty norm name for operator: {op_name}")
            return norm_name

    raise SystemExit(f"operator not found: {op_name}")


def backfill_pr_url(op_name, pr_url, xlsx_path):
    op_name = clean_op(op_name)
    if not op_name:
        raise SystemExit("operator name is empty")
    if not pr_url:
        raise SystemExit("PR URL is empty")
    if not os.path.isfile(xlsx_path):
        raise SystemExit(f"file not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    name_col, norm_col, path_col = find_columns(ws, need_path=True)

    for row in ws.iter_rows(min_row=2):
        raw_name = clean_op(row[name_col - 1].value)
        norm_name = clean_op(row[norm_col - 1].value)
        if op_name in {raw_name, norm_name}:
            row[path_col - 1].value = pr_url
            wb.save(xlsx_path)
            print(f"backfilled {op_name}: {pr_url}")
            return

    raise SystemExit(f"operator not found: {op_name}")


def build_parser():
    parser = argparse.ArgumentParser(description="FlagGems operator registry helper")
    subparsers = parser.add_subparsers(dest="command", required=True)

    lookup = subparsers.add_parser("lookup", help="print normalized operator name")
    lookup.add_argument("operator", help="operator name, with or without aten:: prefix")
    lookup.add_argument("--norm-xlsx", default=DEFAULT_NORM_XLSX, help="path to 规范名.xlsx")

    backfill = subparsers.add_parser("backfill", help="write PR URL to 规范名.xlsx")
    backfill.add_argument("operator", help="operator name, raw or normalized")
    backfill.add_argument("pr_url", help="created PR URL")
    backfill.add_argument("--norm-xlsx", default=DEFAULT_NORM_XLSX, help="path to 规范名.xlsx")

    return parser


def main(argv=None):
    argv = list(sys.argv[1:] if argv is None else argv)
    if argv and argv[0] not in {"lookup", "backfill", "-h", "--help"}:
        argv = ["lookup"] + argv

    args = build_parser().parse_args(argv)
    if args.command == "lookup":
        print(lookup_norm_name(args.operator, args.norm_xlsx))
    elif args.command == "backfill":
        backfill_pr_url(args.operator, args.pr_url, args.norm_xlsx)


if __name__ == "__main__":
    main()
