#!/usr/bin/env python3
"""Resolve one generated FlagGems operator worktree and check upstream conflicts.

Hard constraints in this resolver are identified by stable rule IDs. When a
reviewer-learned naming/worktree/upstream-conflict rule becomes deterministic,
add it here, emit a RESOLVE-H* failure, and update references/hard-constraints.md.
"""

from __future__ import annotations

import argparse
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path


@dataclass(frozen=True)
class Rule:
    rule_id: str
    description: str


RULES = [
    Rule("RESOLVE-H001", "raw operator name is normalized through 规范名.xlsx"),
    Rule("RESOLVE-H002", "normalized operator maps to the canonical module and public op_id"),
    Rule("RESOLVE-H003", "generated worktree exists under the configured worktree root"),
    Rule("RESOLVE-H004", "generated worktree is on the matching gen-* branch"),
    Rule("RESOLVE-H005", "upstream base ref is fetched before conflict checks"),
    Rule("RESOLVE-H006", "upstream does not already contain the target kernel module"),
    Rule("RESOLVE-H007", "upstream does not already contain the target yaml id"),
]

SPECIAL_OPS = {
    "max_pool2d_with_indices_backward": ("max_pool2d_with_indices", "max_pool2d_with_indices_backward"),
    "matmul_layer_norm": ("Matmul_Layer_Norm", "matmul_layernorm"),
    "__and__": ("_and_", "and_op"),
}


def fail(message: str, rule_id: str) -> None:
    known_rule_ids = {rule.rule_id for rule in RULES}
    if rule_id not in known_rule_ids:
        raise AssertionError(f"unknown hard rule id: {rule_id}")
    print(f"ERROR[{rule_id}]: {message}", file=sys.stderr)
    raise SystemExit(1)


def strip_aten(name: object) -> str:
    text = "" if name is None else str(name).strip()
    return text[6:] if text.startswith("aten::") else text


def run_git(worktree: Path, *args: str, check: bool = True) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        ["git", "-C", str(worktree), *args],
        check=check,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )


def resolve_norm_name(raw_op: str, norm_xlsx: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment setup issue
        raise SystemExit("openpyxl is required to read norm-name xlsx files") from exc

    wb = load_workbook(norm_xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = ["" if value is None else str(value).strip() for value in next(rows)]
    except StopIteration:
        fail(f"empty norm-name spreadsheet: {norm_xlsx}", "RESOLVE-H001")

    try:
        op_col = headers.index("算子名")
    except ValueError:
        fail("norm-name spreadsheet must contain a column named 算子名", "RESOLVE-H001")

    norm_cols = [index for index, header in enumerate(headers) if "规范命名" in header]
    if not norm_cols:
        fail("norm-name spreadsheet must contain a column whose header contains 规范命名", "RESOLVE-H001")
    norm_col = norm_cols[0]

    target = strip_aten(raw_op)
    for row in rows:
        name = strip_aten(row[op_col] if op_col < len(row) else None)
        if name == target:
            normalized = "" if norm_col >= len(row) or row[norm_col] is None else str(row[norm_col]).strip()
            if not normalized:
                fail(f"normalized name is empty for {raw_op}", "RESOLVE-H001")
            return normalized

    fail(f"no norm-name match for {raw_op}", "RESOLVE-H001")


def resolve_module_and_id(op: str) -> tuple[str, str]:
    if op in SPECIAL_OPS:
        return SPECIAL_OPS[op]
    return op, op.lstrip("_")


def resolve_worktree(worktree_root: Path, op: str, op_id: str) -> Path:
    candidates = [worktree_root / f"gen-{op}"]
    public_id_candidate = worktree_root / f"gen-{op_id}"
    if public_id_candidate not in candidates:
        candidates.append(public_id_candidate)

    existing = [candidate for candidate in candidates if candidate.is_dir()]
    if not existing:
        fail("generated worktree not found; tried " + ", ".join(str(path) for path in candidates), "RESOLVE-H003")
    return existing[0].resolve()


def check_branch(worktree: Path, op: str, op_id: str) -> str:
    result = run_git(worktree, "rev-parse", "--abbrev-ref", "HEAD")
    branch = result.stdout.strip()
    expected = {f"gen-{op}", f"gen-{op_id}"}
    if branch not in expected:
        fail(f"worktree branch is {branch}, expected one of {sorted(expected)}", "RESOLVE-H004")
    return branch


def fetch_upstream(worktree: Path, upstream: str, base_branch: str) -> str:
    ref = f"refs/remotes/{upstream}/{base_branch}"
    try:
        run_git(worktree, "fetch", upstream, f"{base_branch}:{ref}")
    except subprocess.CalledProcessError as exc:
        fail(f"cannot fetch {upstream}/{base_branch}: {exc.stderr.strip()}", "RESOLVE-H005")
    return ref


def upstream_has_path(worktree: Path, base_ref: str, path: str) -> bool:
    result = run_git(worktree, "cat-file", "-e", f"{base_ref}:{path}", check=False)
    return result.returncode == 0


def upstream_has_yaml_id(worktree: Path, base_ref: str, op_id: str) -> bool:
    result = run_git(worktree, "show", f"{base_ref}:conf/operators.yaml", check=False)
    if result.returncode != 0:
        fail(f"cannot read conf/operators.yaml from {base_ref}", "RESOLVE-H007")
    target_lines = {f"id: {op_id}", f"id: '{op_id}'", f'id: "{op_id}"'}
    for line in result.stdout.splitlines():
        if line.strip() in target_lines:
            return True
    return False


def print_rule_list() -> None:
    for rule in RULES:
        print(f"{rule.rule_id}\t{rule.description}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--raw-op")
    parser.add_argument("--norm-xlsx", type=Path)
    parser.add_argument("--worktree-root", type=Path)
    parser.add_argument("--upstream", default="upstream")
    parser.add_argument("--base-branch", default="master")
    parser.add_argument("--list-rules", action="store_true", help="print hard-rule IDs enforced by this script")
    args = parser.parse_args()

    if args.list_rules:
        return args

    missing = [name for name in ("raw_op", "norm_xlsx", "worktree_root") if getattr(args, name) is None]
    if missing:
        parser.error("missing required arguments unless --list-rules is used: " + ", ".join(f"--{name.replace('_', '-')}" for name in missing))
    return args


def main() -> None:
    args = parse_args()
    if args.list_rules:
        print_rule_list()
        return

    op = resolve_norm_name(args.raw_op, args.norm_xlsx)
    module, op_id = resolve_module_and_id(op)
    worktree = resolve_worktree(args.worktree_root, op, op_id)
    branch = check_branch(worktree, op, op_id)
    base_ref = fetch_upstream(worktree, args.upstream, args.base_branch)

    kernel_path = f"src/flag_gems/ops/{module}.py"
    if upstream_has_path(worktree, base_ref, kernel_path):
        fail(f"upstream already has {kernel_path}", "RESOLVE-H006")
    if upstream_has_yaml_id(worktree, base_ref, op_id):
        fail(f"upstream already has yaml id {op_id}", "RESOLVE-H007")

    print(f"OP={op}")
    print(f"OP_ID={op_id}")
    print(f"MODULE={module}")
    print(f"GEN_WORKTREE={worktree}")
    print(f"GEN_BRANCH={branch}")
    print(f"UPSTREAM_REF={base_ref}")


if __name__ == "__main__":
    main()
